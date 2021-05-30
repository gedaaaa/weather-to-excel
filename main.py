import openpyxl as excel
from openpyxl.styles import PatternFill
import json
import datetime
import asyncio
import aiohttp
import os

input_excel_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'city-list.xlsx')

output_excel_path = os.path.join(os.path.abspath(os.path.dirname(__file__)),  'weather-predict.xlsx')

qweather_key = 'your-key'

location_id_query_url = 'https://geoapi.qweather.com/v2/city/lookup'
weather_predict_24h_url = 'https://devapi.qweather.com/v7/weather/24h'
weather_predict_7d_url = 'https://devapi.qweather.com/v7/weather/7d'

city_id_name_map = {}
global session


async def get(url, params):
    global session
    try:
        response = await session.get(url=url, params=params)
        result = await response.text()
    except:
        print('network exception')
        result = '{"code":"666"}'
    return result


def is_list_all_none(data):
    for item in data:
        if item is not None:
            return False
    return True


def remove_empty_lines():
    active_sheet = excel.load_workbook(input_excel_path).active
    workbook_with_no_empty = excel.Workbook()

    for row in active_sheet.values:
        if not is_list_all_none(row):
            workbook_with_no_empty.active.append(row)

    workbook_with_no_empty.save(input_excel_path)


def read_city_input():
    result = []

    active_sheet = excel.load_workbook(input_excel_path).active

    for row in active_sheet.iter_rows(values_only=True, min_row=2):
        print('read city input: ' + str(row))
        result.append([str(cell) for cell in row])

    return result


async def fetch_location_ids(city_id_and_name_list):
    result = {}
    loop = asyncio.get_event_loop()
    tasks = []

    async def async_execute(city):
        params = {'location': city[2], 'key': qweather_key}
        response = await get(location_id_query_url, params)
        response_data = json.loads(response)
        if response_data['code'] == '200':
            print('location id for city: ' + city_id_name_map.get(city[0], ''))
            # DIDI city id: Location id
            result[city[0]] = response_data.get('location', [])[0].get('id', None)
        else:
            print('Network response with error: ' + response_data['code'])

    for city in city_id_and_name_list:
        tasks.append(loop.create_task(async_execute(city)))

    try:
        await asyncio.wait(tasks)
    finally:
        return result


async def fetch_24h_weather_prediction(city_id_location_id_map):
    result = {}
    loop = asyncio.get_event_loop()
    tasks = []

    async def async_execute(city_id, location_id):
        params = {'location': str(location_id), 'key': qweather_key}
        response = await get(weather_predict_24h_url, params)
        response_data = json.loads(response)
        city_data = []
        if response_data['code'] == '200':
            print('hourly data for city: ' + city_id_name_map.get(city_id, ''))

            for data in response_data.get('hourly', []):
                city_data.append({
                    'fx_time': data.get('fxTime', None),
                    'text': data.get('text', None),
                    'wind': data.get('windScale', None)
                })

            result[city_id] = city_data

        else:
            print('Network response with error: ' + response_data['code'])

    for city_id, location_id in city_id_location_id_map.items():
        tasks.append(loop.create_task(async_execute(city_id, location_id)))

    try:
        await asyncio.wait(tasks)
    finally:
        return result


async def fetch_7d_weather_prediction(city_id_location_id_map):
    result = {}

    loop = asyncio.get_event_loop()
    tasks = []

    async def async_execute(city_id, location_id):
        params = {'location': str(location_id), 'key': qweather_key}
        response = await get(weather_predict_7d_url, params)
        response_data = json.loads(response)
        city_data = []

        if response_data['code'] == '200':
            print('daily data for city: ' + city_id_name_map.get(city_id, ''))

            for data in response_data.get('daily', []):
                city_data.append({
                    'fx_date': data.get('fxDate', None),
                    'text_day': data.get('textDay', None),
                    'text_night': data.get('textNight', None),
                    'wind_day': data.get('windScaleDay', None),
                    'wind_night': data.get('windScaleNight', None)
                })

            result[city_id] = city_data

        else:
            print('Network response with error: ' + response_data['code'])

    for city_id, location_id in city_id_location_id_map.items():
        tasks.append(loop.create_task(async_execute(city_id, location_id)))

    try:
        await asyncio.wait(tasks)
    finally:
        return result


def generate_header():
    head = ['city_id', '省', '城市']
    head.extend(['' for i in range(24 + 7 * 2)])

    now = datetime.datetime.now()
    for i in range(24):
        head[3 + i] = (now + datetime.timedelta(hours=i)).strftime("%m-%d %H:00")
    for i in range(7):
        head[3 + 24 + 2 * i] = (now + datetime.timedelta(days=i)).strftime("%m-%d") + ' 白天'
        head[3 + 24 + 2 * i + 1] = (now + datetime.timedelta(days=i)).strftime("%m-%d") + ' 夜间'
    return head


def color_cells(sheet):
    very_high = PatternFill("solid", fgColor="8E44AD")  # purple
    high = PatternFill("solid", fgColor="E74C3C")  # red
    mid = PatternFill("solid", fgColor="F1C40F")  # yellow
    low = PatternFill("solid", fgColor="3498DB")  # blue
    special = PatternFill("solid", fgColor="E59866")  # orange

    for row in sheet.iter_rows(min_col=4):
        for cell in row:
            if cell.value is None:
                break
            if '雨' in cell.value or '雪' in cell.value:
                cell.fill = low
                if '中' in cell.value:
                    cell.fill = mid
                if '大' in cell.value:
                    cell.fill = high
                if '暴' in cell.value:
                    cell.fill = very_high
                if '雷' in cell.value:
                    cell.fill = special


def generate_city_row(city,city_hourly_predict,city_daily_predict):
    row = []
    row.extend(city)

    row.extend([item['text'] for item in city_hourly_predict.get(city[0], [])])
    for day in city_daily_predict.get(city[0], []):
        row.append(day['text_day'])
        row.append(day['text_night'])

    return row


def generate_report(city_id_name_list, city_hourly_predict, city_daily_predict):
    workbook = excel.Workbook()
    sheet = workbook.active
    sheet.append(generate_header())

    for city in city_id_name_list:
        sheet.append(generate_city_row(city, city_hourly_predict, city_daily_predict))

    color_cells(sheet)

    workbook.save(output_excel_path)
    return


def fill_name_map(city_id_name_list):
    for city in city_id_name_list:
        city_id_name_map[city[0]] = city[2]


async def generate_weather_report():
    global session
    try:
        session = aiohttp.ClientSession()

        print('--- start ---')

        print('--- format input file, sometime very slow ---')
        remove_empty_lines()

        print('--- read input ---')
        city_id_name_list = read_city_input()
        fill_name_map(city_id_name_list)

        print('--- fetch result for {} cities ---'.format(len(city_id_name_list)))
        city_id_location_id_map = await fetch_location_ids(city_id_name_list)
        city_hourly_predict = await fetch_24h_weather_prediction(city_id_location_id_map)
        city_daily_predict = await fetch_7d_weather_prediction(city_id_location_id_map)

        print('--- generating excel output ---')
        generate_report(city_id_name_list, city_hourly_predict, city_daily_predict)
    finally:
        await session.close()
    print('--- FINISH ---')

if __name__ == '__main__':
    asyncio.run(generate_weather_report())

